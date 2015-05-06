#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Created on Tue Dec 02 12:53:35 2014

@author: Wiebke.Toussaint
"""

import pandas as pd
from pandas import ExcelWriter
from datetime import date, timedelta
from openpyxl.reader.excel import load_workbook
import MyFunx, AllData

today = date.today()

lastmonth = today.month - 3
nextmonth = today.month + 1

Visibility = AllData.InboundData(lastmonth, nextmonth, today)

Samples = pd.ExcelFile(u'05_Samples\\SampleTrack.xlsx').parse(u'Master', skiprows = 0, index = None, encoding='utf-8')
Samples = Samples[[u'SKU',u'03_SampleRoom_TO_studio']]
Samples.loc[Samples[u'03_SampleRoom_TO_studio'].notnull(),u'SampleCount'] = 1

Vis = pd.merge(Visibility, Samples, on = u'SKU', how = u'left', sort = False )

V = Vis.sort(['Date received','Date booked','POs'], inplace = False, na_position = 'first')
V = V[['GLYear','GLMonth','GLDay','Buyer', 'UnitCost','TotalUnits','TotalCost','SKU','SimpleName','ProcurementStatus','Category','Supplier','DeliveryDue','POs','BP Qty','Ref','Status','Date booked','Partial delivery','Date received','LastQCed','Qty Counted','Qty Damaged','SampleCount','Oversupply','OTDLastReceived','Qty Received','Qty PutAway','ActualGoLiveDate']]
 
today = date.today()

#==============================================================================
# Generate MerchTrack Output Data
#==============================================================================
def format():
    worksheet.set_column('A:C', 8 )
    worksheet.set_column('D:D', 12 )
    worksheet.set_column('E:G', 10 )
    worksheet.set_column('H:H', 16 )
    worksheet.set_column('I:I', 32 )
    worksheet.set_column('J:M', 20 )
    worksheet.set_column('N:O', 8 )
    worksheet.set_column('P:U', 18 )
    worksheet.set_column('V:Y', 12 )
    worksheet.set_column('Z:Z', 18 )
    worksheet.set_column('AA:AB', 12 )
    worksheet.set_column('AC:AC', 18 )

writer1 = ExcelWriter('04_Visibility\\Visibility ' + str(today) + '.xlsx')
V.to_excel(writer1, 'MASTER', index = False, encoding = 'utf-8' )   
workbook = writer1.book
worksheet = writer1.sheets['MASTER']
format()
Buyers = pd.Series(V[u'Buyer'].unique())
for b in Buyers:
    DataName = V[V['Buyer']==b]
    DataName.to_excel(writer1, b , index = False )   
    workbook = writer1.book
    worksheet = writer1.sheets[b]
    format()
writer1.save()

doc_name = u'Visibility Report '
part = u'04_Visibility\\Visibility ' + str(today) + '.xlsx'
message = u'Visibility Report' + str(today)
maillist = "MailList_Merch.txt"

MyFunx.send_message(doc_name, message, part, maillist)

#==============================================================================
# Generate WHTrack Output Data
#==============================================================================
WHTrack = V[['SKU','SimpleName','Category','Supplier','POs','DeliveryDue','Date booked','Date received','LastQCed',\
'OTDLastReceived','UnitCost','TotalUnits','TotalCost','BP Qty','Qty Counted','Qty Damaged','Qty Received','Qty PutAway',\
'Partial delivery','Buyer','ProcurementStatus','Status','Ref','GLMonth']]
WHTrack = WHTrack.sort(columns = ['Date received','POs','LastQCed'], ascending = True, na_position = 'last', inplace = False)
WHTrack.name = 'WHTrack'

Backlog = WHTrack.dropna(subset = ['Date received'], inplace = False)
Backlog = Backlog[(Backlog['LastQCed'].isnull()==True) & (Backlog['OTDLastReceived'].isnull()==True)]
Backlog.name = 'Backlog'

def format1():
    worksheet.set_column('A:A', 15 )
    worksheet.set_column('B:B', 40 )
    worksheet.set_column('D:C', 15 )
    worksheet.set_column('E:E', 10 )
    worksheet.set_column('F:J', 18 )
    worksheet.set_column('K:R', 10 )
    worksheet.set_column('S:W', 20 )
    worksheet.set_column('X:X', 6 )

writer2 = ExcelWriter('04_Visibility\\WHTrack ' + str(today) + '.xlsx')
Track = [WHTrack,Backlog]
for t in Track:
    t.to_excel(writer2, t.name , index = False )   
    workbook = writer2.book
    worksheet = writer2.sheets[t.name]
    format1()  
writer2.save()
   
doc_name = 'WH Stock Track Report '
part = '04_Visibility\\WHTrack ' + str(today) + '.xlsx'
message = 'Spree Stock Tracking on ' + str(today)
maillist = "MailList_WH.txt"

MyFunx.send_message(doc_name, message, part, maillist)

#==============================================================================
# Generate ProductTrack QuickStats
#==============================================================================

#ProductTrack by Config 

V = Vis[(Vis['Ref'].str.contains(u"sample|Sample|SAMPLE|samples|Samples|OS|Os|OVERSUPPLY|fraud|not OK")==False) | (Vis['Ref'].isnull()==True)] 

def iso_year_start(iso_year):
    "The gregorian calendar date of the first day of the given ISO year"
    fourth_jan = date(iso_year, 1, 4)
    delta = timedelta(fourth_jan.isoweekday()-1)
    return fourth_jan - delta 

def iso_to_gregorian(iso_year, iso_week, iso_day):
    "Gregorian calendar date for the given ISO year, week and day"
    year_start = iso_year_start(iso_year)
    return year_start + timedelta(days=iso_day-1, weeks=iso_week-1)
    
V['GLiso'] = V['GLDate'].apply(lambda x: list(x.isocalendar())[:-1] + [3])
V['GLgreg'] = V['GLiso'].apply(lambda x: iso_to_gregorian(x[0],x[1],x[2]))

VConf = Visibility.drop_duplicates(subset = ['ConfigSKU','LastQCed'])
VConf.sort(columns = 'LastQCed', na_position = 'last', inplace = True)
VConf['duplicate'] = VConf.duplicated(subset = 'POs', take_last = False)
VConf.loc[VConf['LastQCed'].notnull(), 'atWH'] = 1
VConf = VConf.drop_duplicates(subset = ['POs','atWH'])

GLMonth = VConf.groupby('GLMonth')
MonthTrack = GLMonth.apply(lambda x : pd.Series(dict(
ConfigTotal = len(x.loc[x.duplicate==False]),
ConfigonBP = len(x.loc[(x.duplicate==False) & x.POs.notnull()]),
ConfigBooked = len(x.loc[(x.duplicate==False) & x['Date booked'].notnull()]),
ConfigRec = len(x.loc[(x.duplicate==False) & x['Date received'].notnull()]),
ConfigParDel = len(x.loc[(x.duplicate==True) & x['LastQCed'].isnull()]),
ConfigQCd = len(x.loc[(x.duplicate==False) & x['LastQCed'].notnull()]),
ConfigLive = len(x.loc[(x.duplicate==False) & x['ActualGoLiveDate'].notnull()]),
ConfigNotLive = len(x.loc[(x.duplicate==False) & x['ActualGoLiveDate'].isnull()]))))
MonthTrack = MonthTrack.rename(columns={'ConfigTotal':'Planned','ConfigonBP':'OnBP','ConfigRec':'Received','ConfigQCd':'QCed','ConfigBooked':'Booked','ConfigLive':'Live','ConfigNotLive':'Not Live','ConfigParDel':'Partial delivery'})
MonthTrack = MonthTrack[['Planned','OnBP','Booked','Received','QCed','Partial delivery','Live','Not Live']]

UnitsCount = V.groupby('GLMonth').agg({'TotalUnits':'sum','Qty Counted':'sum','SampleCount':'sum','Qty Damaged':'sum','Qty Received':'sum','Qty PutAway':'sum'})
UnitsCount.sort_index(ascending = True, inplace = True)
UnitsCount = UnitsCount[['TotalUnits','Qty Counted','SampleCount','Qty Damaged','Qty Received','Qty PutAway']]
UnitsCount.rename(columns={'TotalUnits':'Units Planned','Qty Counted':'Units QCed','Qty Received':'Units taken in by OTD','Qty PutAway':'Units in OTD WH'}, inplace = True)
UnitsCount.name = "Units Count"

#Stats by Purchase Order
POCount = V.dropna(subset = ['POs'])
POCount = POCount.drop_duplicates(subset = ['POs','LastQCed'])
POCount.sort(columns = 'LastQCed', na_position = 'last', inplace = True)
POCount['duplicate'] = POCount.duplicated(subset = 'POs', take_last = False)
POCount.loc[POCount['LastQCed'].notnull(), 'atWH'] = 1
POCount = POCount.drop_duplicates(subset = ['POs','atWH'])

POC = POCount.groupby('GLMonth')
POC = POC.apply(lambda x : pd.Series(dict(
POTotal = len(x.loc[x.duplicate==False]), 
POBooked = len(x.loc[(x.duplicate==False) & x['Date booked'].notnull()]), 
POReceived = len(x.loc[(x.duplicate==False) & x['Date received'].notnull()]),
PONotRec = len(x.loc[(x.duplicate==False) & x['Date received'].isnull()]),
Partial = len(x.loc[(x.duplicate==True) & x['LastQCed'].isnull()]), 
LastQCed = len(x.loc[(x.duplicate==False) & x['LastQCed'].notnull()]), 
OTDLastReceived = len(x.loc[(x.duplicate==False) & x['OTDLastReceived'].notnull()]))))

POC = POC[['POTotal', 'POBooked', 'POReceived','PONotRec','Partial','LastQCed','OTDLastReceived']]
POC.rename(columns={'POTotal':'POs on BP','POBooked':'POs booked','POReceived':'POs received UNCHECKED','Partial':'Partial delivery','LastQCed':'POs QCed','OTDLastReceived':'POs in WH'}, inplace = True)
POC = (POC.T/list(POC['POs on BP'])).T
POC.name = "PO Count"

#SKUs not processed
V['TotalCost'] = V['UnitCost']*V['TotalUnits']
V['NoBP'] = V['Status'].isnull() #SKUs not on Brightpearl
V['NBND'] = V['Date received'].isnull() & V['Date booked'].isnull() #SKUs not booked not delivered
V['ND'] = V['Date received'].isnull() #SKUs not delivered
V['Partial'] = V['LastQCed'].isnull() #SKUs not QCed
V['NOTD'] = V['OTDLastReceived'].isnull() #SKUs not received by OTD

#Cost of SKUs not processed / WorkingCapital
NoBP = V[V['NoBP']==True].groupby(V['GLMonth']).sum()['TotalCost'] #Cost of SKUs not on BP / month
NBND = V[V['NBND']==True].groupby(V['GLMonth']).sum()['TotalCost'] - NoBP #Cost of SKUs not booked not delivered
BND = V[V['ND']==True].groupby(V['GLMonth']).sum()['TotalCost'] - (NoBP + NBND) #Cost of SKUs not delivered / month
Partial = V[V['Partial']==True].groupby(V['GLMonth']).sum()['TotalCost'] - (NoBP + NBND + BND)  #Cost of SKUs not QCed
#NOTD = V[V['NOTD']==True].groupby(V['GLMonth']).sum()['TotalCost'] - (NoBP + NBND + BND + Partial)#Cost of SKUs not received by OTD

idx = ['Not on Brightpearl', 'Not Booked Not Delivered', 'Booked Not Delivered', 'Partial Delivery']
WorkingCapital = pd.DataFrame(data = [NoBP, NBND, BND, Partial], index = idx).T
WorkingCapital.applymap(lambda x: "R{:.8n}".format(x))
WorkingCapital.name = "Working Capital"

writer3 = ExcelWriter('04_Visibility\\ProductTrack QuickStats ' + str(today) + '.xlsx')
SimplesCount.to_excel(writer3, 'Sheet1', startrow = 3)
UnitsCount.to_excel(writer3, 'Sheet1', startrow = 10)
POC.to_excel(writer3, 'Sheet1', startrow = 17)
WorkingCapital.to_excel(writer3, 'Sheet1', startrow = 24)
workbook = writer3.book
#format workbook
title = workbook.add_format({'bold':True, 'size':14})
header = workbook.add_format({'size':12, 'underline':True, 'font_color':'green'})
worksheet = writer3.sheets['Sheet1']
worksheet.write('A1','Spree Stock Tracking Statistics ' + str(today), title)
worksheet.write('A3','Simples Count (% of Simples Planned)', header)
worksheet.write('A11','Units Count', header)
worksheet.write('A18','PO Count', header)
worksheet.write('A25','Working Capital (ZAR loss due to status not achieved)', header)
worksheet.set_column('A:A', 8 )
worksheet.set_column('B:K', 18)
writer3.save()

#format QuickStats with openpyxl
wb = load_workbook('04_Visibility\\ProductTrack QuickStats ' + str(today) + '.xlsx')
ws = wb.worksheets[0]

cellsA = [ws['E4'],ws['E11'],ws['G4'],ws['D18'],ws['C25'],ws['D25']]
for cell in cellsA:
    cell.style.alignment.wrap_text = True 
    
cellsB = ws['B20':'F24']
for row in cellsB:
    for cell in row:
        cell.style.number_format.format_code = '0%'
        cell.style.alignment.horizontal = 'center'

cellsC = ws['B27':'F31']
for row in cellsC:
    for cell in row:
        cell.style.number_format.format_code = '"R "#,##0.00'
        
wb.save('04_Visibility\\ProductTrack QuickStats ' + str(today) + '.xlsx')

doc_name = 'ProductTrack QuickStats '
part = '04_Visibility\\ProductTrack QuickStats ' + str(today) + '.xlsx'
message = 'Where is my stock? Quick Stats to monitor production progress'
maillist = "MailList_QS.txt"

MyFunx.send_message(doc_name, message, part, maillist)

#==============================================================================
# Production Track Weekly Stats
#==============================================================================

GLTrack = VConf.groupby(['GLgreg'])
ConfigTrack = GLTrack.agg({'ConfigSKU':'count','POs':'count','Date booked':'count','Date received':'count','LastQCed':'count','ActualGoLiveDate':'count'})
ConfigTrack = ConfigTrack.rename(columns={'ConfigSKU':'Planned','POs':'OnBP','Date received':'Received','LastQCed':'QCed','Date booked':'Booked','ActualGoLiveDate':'Live'})
ConfigTrack = ConfigTrack[['Planned','OnBP','Booked','Received','QCed','Live']]

POTrack = V.dropna(subset = ['POs'])
POTrack = POTrack.drop_duplicates(subset = ['POs','LastQCed'])
POTrack.sort(columns = 'LastQCed', na_position = 'last', inplace = True)
POTrack['duplicate'] = POTrack.duplicated(subset = 'POs', take_last = False)

POsum = POTrack.groupby(['GLMonth','GLDay'])
POsum = POsum.apply(lambda x : pd.Series(dict(
POsum_Total = len(x.loc[x.duplicate==False]),
POsum_Draft = len(x.loc[(x.duplicate==False) & (x['Status'] == 'Draft PO')]),
sumNOT_Received = len(x.loc[(x.duplicate==False) & x['Date received'].isnull()]),
POsum_Received = len(x.loc[(x.duplicate==False) & x['Date received'].notnull()]))))
POsum = POsum[['POsum_Total','POsum_Draft','POsum_Received','sumNOT_Received']]
POsum.name = "PO Track Summary"

POT = POTrack.groupby(['Buyer','GLMonth','GLDay'])
POT = POT.apply(lambda x : pd.Series(dict(
PO_Total = len(x.loc[x.duplicate==False]),
PO_Draft = len(x.loc[(x.duplicate==False) & (x['Status'] == 'Draft PO')]),
NOT_Received = len(x.loc[(x.duplicate==False) & x['Date received'].isnull()]),
PO_Received = len(x.loc[(x.duplicate==False) & x['Date received'].notnull()]))))
POT = POT[['PO_Total','PO_Draft','PO_Received','NOT_Received']]
POT.name = "PO Track"

doc_name = 'GLTrack QuickStats '
part = '04_Visibility\\GoLiveTrack QuickStats ' + str(today) + '.xlsx'
message = 'Quick Stats to monitor Purchase Orders received to meet Go Live Dates'
maillist = "MailList_Prod.txt"

writer4 = ExcelWriter(part)
POsum.to_excel(writer4, 'GLperPO Summary', startrow = 2)
POT.to_excel(writer4, 'GLperPO per Buyer', startrow = 2)
workbook = writer4.book
#format workbook
title = workbook.add_format({'bold':True, 'size':14})
header = workbook.add_format({'size':12, 'underline':True, 'font_color':'green'})
worksheet = writer4.sheets['GLTrack per Buyer']
worksheet.write('A1','Spree Inbound POs to make Go Live Statistics ' + str(today), title)
worksheet.set_column('A:K', 12 )

ws = writer4.sheets['GLTrack Summary']
ws.write('A1','Spree Inbound POs to make Go Live Statistics ' + str(today), title)
ws.set_column('A:K', 12 )
writer4.save()

MyFunx.send_message(doc_name, message, part, maillist)